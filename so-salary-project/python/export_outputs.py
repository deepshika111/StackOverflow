from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PARQUET_PATH = PROJECT_ROOT / "data" / "processed" / "survey_clean.parquet"
OUTPUTS_DIR = PROJECT_ROOT / "outputs"


def prettify_skill_name(column_name: str) -> str:
    return column_name.split("_", 1)[1].replace("_", " ").title()


def explode_roles(df: pd.DataFrame) -> pd.DataFrame:
    roles = (
        df[["DevType", "ConvertedCompYearly", "exp_bucket", "country_group"]]
        .dropna(subset=["DevType", "exp_bucket", "country_group"])
        .assign(role=lambda frame: frame["DevType"].str.split(";"))
        .explode("role")
    )
    roles["role"] = roles["role"].str.strip()
    return roles[roles["role"].ne("")]


def build_skill_premiums(df: pd.DataFrame) -> pd.DataFrame:
    skill_columns = [
        column
        for column in df.columns
        if column.startswith(("lang_", "db_", "plat_"))
    ]
    data_roles = df[df["is_data_role"] == 1].copy()

    rows: list[dict[str, float | int | str]] = []
    for skill in skill_columns:
        with_skill = data_roles.loc[data_roles[skill] == 1, "ConvertedCompYearly"]
        without_skill = data_roles.loc[data_roles[skill] == 0, "ConvertedCompYearly"]
        if with_skill.empty or without_skill.empty:
            continue

        with_mean = float(with_skill.mean())
        without_mean = float(without_skill.mean())
        rows.append(
            {
                "skill": skill,
                "skill_label": prettify_skill_name(skill),
                "skill_family": skill.split("_", 1)[0],
                "n_with": int(with_skill.shape[0]),
                "n_without": int(without_skill.shape[0]),
                "prevalence_pct": float(data_roles[skill].mean() * 100),
                "with_skill": with_mean,
                "without_skill": without_mean,
                "premium_pct": float((with_mean / without_mean - 1) * 100),
            }
        )

    return pd.DataFrame(rows).sort_values("premium_pct", ascending=False)


def build_benchmarking_table(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    role_rows = explode_roles(df)
    benchmark_long = (
        role_rows.groupby(["role", "country_group", "exp_bucket"])["ConvertedCompYearly"]
        .agg(
            median=lambda salary: salary.quantile(0.5),
            p75=lambda salary: salary.quantile(0.75),
            n="count",
        )
        .reset_index()
        .query("n >= 10")
        .sort_values(["role", "country_group", "exp_bucket"])
    )

    benchmark_wide = benchmark_long.pivot_table(
        values=["median", "p75", "n"],
        index=["role", "country_group"],
        columns="exp_bucket",
    )

    benchmark_wide.columns = [
        f"{metric}_{bucket}" for metric, bucket in benchmark_wide.columns.to_flat_index()
    ]
    benchmark_wide = benchmark_wide.reset_index()
    return benchmark_long, benchmark_wide


def format_benchmark_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.auto_filter.ref = worksheet.dimensions

    if worksheet.max_row >= 2 and worksheet.max_column >= 3:
        worksheet.conditional_formatting.add(
            f"C2:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}",
            ColorScaleRule(
                start_type="percentile",
                start_value=10,
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="percentile",
                end_value=90,
                end_color="63BE7B",
            ),
        )

    for column_cells in worksheet.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 24)

    workbook.save(path)


def main() -> None:
    df = pd.read_parquet(PARQUET_PATH)
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    feature_matrix_path = OUTPUTS_DIR / "feature_matrix.csv"
    skill_premiums_path = OUTPUTS_DIR / "skill_premiums.csv"
    benchmarking_csv_path = OUTPUTS_DIR / "benchmarking_table.csv"
    benchmarking_xlsx_path = OUTPUTS_DIR / "benchmarking_table.xlsx"

    df.to_csv(feature_matrix_path, index=False)

    skill_premiums = build_skill_premiums(df)
    skill_premiums.to_csv(skill_premiums_path, index=False)

    benchmark_long, benchmark_wide = build_benchmarking_table(df)
    benchmark_long.to_csv(benchmarking_csv_path, index=False)
    benchmark_wide.to_excel(benchmarking_xlsx_path, index=False)
    format_benchmark_workbook(benchmarking_xlsx_path)

    print(f"Wrote {feature_matrix_path}")
    print(f"Wrote {skill_premiums_path}")
    print(f"Wrote {benchmarking_csv_path}")
    print(f"Wrote {benchmarking_xlsx_path}")


if __name__ == "__main__":
    main()
